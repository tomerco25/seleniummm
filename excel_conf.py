from openpyxl import load_workbook
#read from excel
def activetests():
    counter=1
    truetests=[]
    wb = load_workbook(filename = 'C:/1/tests.xlsx')
    sheet_ranges = wb['Sheet1']
    while (sheet_ranges['A'+str(counter)+''].value != None):
        if sheet_ranges['C'+str(counter)+''].value==1:
            truetests.append(sheet_ranges['A'+str(counter)+''].value)
            counter = int(counter) + 1
        else:
            counter = int(counter) + 1
    return truetests
 # #write to excel
# ws=wb.active
# ws['D2']=432432
# wb.save('C:/excel/1.xlsx')
#
# #read units id from excel and save
# count = '1'
# while (sheet_ranges['A'+str(count)+''].value != None):
#    wb = load_workbook(filename = 'C:/excel/1.xlsx')
#    sheet_ranges = wb['Sheet1']
#    count=str(count)
#    unitid=sheet_ranges['A'+count+''].value
#    print unitid
#    count = int(count) + 1